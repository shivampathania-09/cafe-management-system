from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.db.models import Sum, F, Q, Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib import messages
from django.contrib.auth import login
from django.utils.decorators import method_decorator
from .models import MenuItem, Order, OrderItem, Activity, Notification
from .forms import MenuItemForm, UserRegistrationForm, RoleSelectionLoginForm
from .decorators import admin_required, staff_required
from decimal import Decimal
import xhtml2pdf.pisa as pisa
from django.utils import timezone
from datetime import timedelta
import json
import csv
import openpyxl

def register(request):
    if request.user.is_authenticated:
        return redirect('cafe:dashboard')
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.role = 'staff'
            user.save()
            messages.success(request, 'Registration successful! Please login.')
            return redirect('cafe:login')
    else:
        form = UserRegistrationForm()
    return render(request, 'cafe/register.html', {'form': form})

def custom_login_view(request):
    if request.user.is_authenticated:
        return redirect('cafe:dashboard')
    
    if request.method == 'POST':
        form = RoleSelectionLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")
            # The next parameter logic could be handled here, but redirecting to dashboard is safe
            return redirect('cafe:dashboard')
    else:
        form = RoleSelectionLoginForm()
        
    return render(request, 'cafe/login.html', {'form': form})

@login_required
def dashboard(request):
    if request.user.role == 'admin' or request.user.is_superuser:
        return redirect('cafe:admin_dashboard')
    return redirect('cafe:staff_dashboard')

@admin_required
def admin_dashboard(request):
    total_orders = Order.objects.count()
    total_menu_items = MenuItem.objects.count()
    recent_orders = Order.objects.order_by('-created_at')[:5]
    
    most_ordered = OrderItem.objects.values('menu_item__name').annotate(
        total_qty=Sum('quantity')
    ).order_by('-total_qty').first()

    pending_orders_count = Order.objects.filter(status='pending').count()
    
    # Top 5 Best Selling Items for Chart
    top_items = OrderItem.objects.values('menu_item__name').annotate(
        total_qty=Sum('quantity')
    ).order_by('-total_qty')[:5]
    best_selling_labels = [item['menu_item__name'] for item in top_items]
    best_selling_data = [item['total_qty'] for item in top_items]
    
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    # Today's Revenue
    today_orders = Order.objects.filter(created_at__date=today)
    today_revenue = today_orders.aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')

    # Weekly Revenue
    weekly_orders = Order.objects.filter(created_at__date__gte=seven_days_ago)
    weekly_revenue = weekly_orders.aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')

    # Chart Data Preparation
    dates = [(today - timedelta(days=i)) for i in range(6, -1, -1)]
    date_labels = [d.strftime('%b %d') for d in dates]
    revenue_data = []
    order_data = []

    for d in dates:
        daily_orders = Order.objects.filter(created_at__date=d)
        daily_rev = daily_orders.aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')
        daily_count = daily_orders.count()
        revenue_data.append(float(daily_rev))
        order_data.append(daily_count)

    total_low_stock = MenuItem.objects.filter(stock_quantity__gt=0, stock_quantity__lte=F('low_stock_threshold')).count()
    total_out_of_stock = MenuItem.objects.filter(stock_quantity=0).count()
    low_stock_items = MenuItem.objects.filter(stock_quantity__gt=0, stock_quantity__lte=F('low_stock_threshold'))

    recent_activities = Activity.objects.order_by('-created_at')[:10]

    context = {
        'total_orders': total_orders,
        'today_revenue': today_revenue,
        'weekly_revenue': weekly_revenue,
        'recent_orders': recent_orders,
        'most_ordered_item': most_ordered['menu_item__name'] if most_ordered else 'None yet',
        'pending_orders_count': pending_orders_count,
        'date_labels': json.dumps(date_labels),
        'revenue_data': json.dumps(revenue_data),
        'order_data': json.dumps(order_data),
        'best_selling_labels': json.dumps(best_selling_labels),
        'best_selling_data': json.dumps(best_selling_data),
        'total_low_stock': total_low_stock,
        'total_out_of_stock': total_out_of_stock,
        'low_stock_items': low_stock_items,
        'recent_activities': recent_activities,
    }
    return render(request, 'cafe/admin_dashboard.html', context)

@staff_required
def staff_dashboard(request):
    active_orders = Order.objects.filter(status__in=['pending', 'preparing', 'ready']).order_by('-created_at')
    recent_activities = Activity.objects.filter(activity_type__in=['order_created', 'order_status']).order_by('-created_at')[:10]
    
    context = {
        'active_orders': active_orders,
        'recent_activities': recent_activities,
    }
    return render(request, 'cafe/staff_dashboard.html', context)

@method_decorator(admin_required, name='dispatch')
class MenuListView(LoginRequiredMixin, ListView):
    model = MenuItem
    template_name = 'cafe/menu_list.html'
    context_object_name = 'items'

@method_decorator(admin_required, name='dispatch')
class MenuCreateView(LoginRequiredMixin, CreateView):
    model = MenuItem
    form_class = MenuItemForm
    template_name = 'cafe/menu_form.html'
    success_url = reverse_lazy('cafe:menu_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        Activity.objects.create(message=f"{self.object.name} added to menu", activity_type='menu_added')
        messages.success(self.request, f"Menu item '{self.object.name}' added successfully.")
        return response

@method_decorator(admin_required, name='dispatch')
class MenuUpdateView(LoginRequiredMixin, UpdateView):
    model = MenuItem
    form_class = MenuItemForm
    template_name = 'cafe/menu_form.html'
    success_url = reverse_lazy('cafe:menu_list')

    def form_valid(self, form):
        response = super().form_valid(form)
        Activity.objects.create(message=f"{self.object.name} updated", activity_type='menu_updated')
        messages.success(self.request, f"Menu item '{self.object.name}' updated successfully.")
        return response

@method_decorator(admin_required, name='dispatch')
class MenuDeleteView(LoginRequiredMixin, DeleteView):
    model = MenuItem
    template_name = 'cafe/menu_confirm_delete.html'
    success_url = reverse_lazy('cafe:menu_list')

    def form_valid(self, form):
        item_name = self.object.name
        response = super().form_valid(form)
        Activity.objects.create(message=f"{item_name} deleted from menu", activity_type='menu_deleted')
        messages.success(self.request, f"Menu item '{item_name}' deleted successfully.")
        return response

# --- Order & Billing Views ---
@staff_required
def order_create(request):
    if request.method == 'POST':
        table_number = request.POST.get('table_number', '')
        order = Order.objects.create(user=request.user, table_number=table_number)
        
        subtotal = Decimal('0.00')
        menu_items = MenuItem.objects.filter(is_available=True)
        
        for item in menu_items:
            quantity_str = request.POST.get(f'quantity_{item.id}', '0')
            try:
                quantity = int(quantity_str)
                if quantity > 0:
                    if quantity <= item.stock_quantity:
                        OrderItem.objects.create(order=order, menu_item=item, quantity=quantity)
                        subtotal += item.price * quantity
                        
                        item.stock_quantity -= quantity
                        if item.stock_quantity == 0:
                            item.is_available = False
                        item.save()
                        Activity.objects.create(message=f"{item.name} stock updated (-{quantity})", activity_type='inventory_updated')
                        
                        if item.stock_quantity > 0 and item.stock_quantity <= item.low_stock_threshold:
                            messages.warning(request, f"Low stock alert: '{item.name}' has only {item.stock_quantity} left.")
                            Notification.objects.create(
                                notification_type='stock',
                                message=f"Low stock alert: '{item.name}' has only {item.stock_quantity} left."
                            )
                        elif item.stock_quantity == 0:
                            messages.error(request, f"Alert: '{item.name}' is now out of stock!")
                            Notification.objects.create(
                                notification_type='stock',
                                message=f"Alert: '{item.name}' is now out of stock!"
                            )
                    else:
                        messages.error(request, f"Cannot order {quantity} of '{item.name}'. Only {item.stock_quantity} available.")
            except ValueError:
                pass
                
        gst_amount = subtotal * Decimal('0.18')
        total_price = subtotal + gst_amount
        
        order.subtotal = subtotal
        order.gst_amount = gst_amount
        order.total_price = total_price
        order.save()
        Activity.objects.create(message=f"Order #{order.id} created", activity_type='order_created')
        
        messages.success(request, f"Order #{order.id} created successfully.")
        return redirect('cafe:order_list')
        
    items = MenuItem.objects.filter(is_available=True)
    return render(request, 'cafe/order_form.html', {'items': items})

@method_decorator(staff_required, name='dispatch')
class OrderListView(LoginRequiredMixin, ListView):
    model = Order
    template_name = 'cafe/order_list.html'
    context_object_name = 'orders'
    
    def get_queryset(self):
        queryset = super().get_queryset().order_by('-created_at')
        
        date = self.request.GET.get('date')
        table_number = self.request.GET.get('table_number')
        status = self.request.GET.get('status')
        
        if date:
            queryset = queryset.filter(created_at__date=date)
        if table_number:
            queryset = queryset.filter(table_number__icontains=table_number)
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_date'] = self.request.GET.get('date', '')
        context['current_table'] = self.request.GET.get('table_number', '')
        context['current_status'] = self.request.GET.get('status', '')
        return context

@staff_required
def order_invoice_pdf(request, order_id):
    order = Order.objects.get(id=order_id)
    template_path = 'cafe/invoice_pdf.html'
    context = {'order': order}
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{order.id}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@staff_required
def order_update_status(request, order_id):
    if request.method == 'POST':
        order = Order.objects.get(id=order_id)
        new_status = request.POST.get('status')
        if new_status in dict(Order.STATUS_CHOICES).keys():
            order.status = new_status
            order.save()
            Activity.objects.create(message=f"Order #{order.id} status updated to {new_status}", activity_type='order_status')
            status_display = dict(Order.STATUS_CHOICES)[new_status]
            messages.success(request, f"Order #{order.id} status updated to {status_display}.")
            Notification.objects.create(
                user=order.user,
                notification_type='order',
                message=f"Order #{order.id} status updated to {status_display}."
            )
    return redirect('cafe:order_list')

@login_required
def notifications_list(request):
    qs = Notification.objects.filter(user=request.user) | Notification.objects.filter(user__isnull=True)
    notifications = qs.order_by('-created_at')
    return render(request, 'cafe/notifications.html', {'notifications': notifications})

@login_required
def mark_notification_read(request, notification_id):
    if request.method == 'POST':
        try:
            notification = Notification.objects.get(id=notification_id)
            # Basic permission check
            if notification.user == request.user or notification.user is None:
                notification.is_read = True
                notification.save()
        except Notification.DoesNotExist:
            pass
    return redirect('cafe:notifications_list')

# --- Export Views ---
@admin_required
def export_orders(request):
    format = request.GET.get('format', 'csv')
    orders = Order.objects.all().order_by('-created_at')
    
    if format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="orders_report.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(['Order ID', 'Date', 'Customer', 'Table', 'Status', 'Subtotal', 'GST', 'Total'])
        for order in orders:
            ws.append([
                order.id, 
                order.created_at.strftime("%Y-%m-%d %H:%M"), 
                order.user.username, 
                order.table_number, 
                order.get_status_display(),
                float(order.subtotal),
                float(order.gst_amount),
                float(order.total_price)
            ])
        wb.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="orders_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Order ID', 'Date', 'Customer', 'Table', 'Status', 'Subtotal', 'GST', 'Total'])
        for order in orders:
            writer.writerow([
                order.id, 
                order.created_at.strftime("%Y-%m-%d %H:%M"), 
                order.user.username, 
                order.table_number, 
                order.get_status_display(),
                order.subtotal,
                order.gst_amount,
                order.total_price
            ])
        return response

@admin_required
def export_inventory(request):
    format = request.GET.get('format', 'csv')
    items = MenuItem.objects.all().order_by('name')
    
    if format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.append(['Item Name', 'Price', 'Available', 'Stock Quantity', 'Low Stock Threshold'])
        for item in items:
            ws.append([
                item.name, 
                float(item.price), 
                'Yes' if item.is_available else 'No', 
                item.stock_quantity, 
                item.low_stock_threshold
            ])
        wb.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Item Name', 'Price', 'Available', 'Stock Quantity', 'Low Stock Threshold'])
        for item in items:
            writer.writerow([
                item.name, 
                item.price, 
                'Yes' if item.is_available else 'No', 
                item.stock_quantity, 
                item.low_stock_threshold
            ])
        return response

@admin_required
def export_sales(request):
    format = request.GET.get('format', 'csv')
    
    sales_data = Order.objects.annotate(date=TruncDate('created_at')).values('date').annotate(
        total_orders=Count('id'),
        daily_revenue=Sum('total_price')
    ).order_by('-date')

    if format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(['Date', 'Total Orders', 'Daily Revenue'])
        for sale in sales_data:
            ws.append([
                sale['date'].strftime("%Y-%m-%d"), 
                sale['total_orders'], 
                float(sale['daily_revenue'] or 0)
            ])
        wb.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sales_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Date', 'Total Orders', 'Daily Revenue'])
        for sale in sales_data:
            writer.writerow([
                sale['date'].strftime("%Y-%m-%d"), 
                sale['total_orders'], 
                sale['daily_revenue'] or 0
            ])
        return response
